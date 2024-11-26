import os
import pytesseract
from PyPDF2 import PdfReader
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import json

# Directory to process files
UPLOAD_FOLDER = r"F:\repogit\X-seller-8\frontend\public\uploads"
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpeg', 'jpg', 'txt', 'xlsx', 'csv'}
CENTRAL_SPREADSHEET = r"F:\repogit\X-seller-8\backend\central_spreadsheet.xlsx"

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Helper function to check allowed file types
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Functions to process different file types
def process_pdfs(file_path):
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text

def process_images(file_path):
    return pytesseract.image_to_string(file_path)

def process_excel(file_path):
    df = pd.read_excel(file_path)
    return df.to_dict(orient="records")

# Function to create the central spreadsheet if it doesn't exist
def create_central_spreadsheet():
    if not os.path.exists(CENTRAL_SPREADSHEET):
        wb = Workbook()
        wb.create_sheet("Monthly Costs")
        wb.create_sheet("Sales Data")
        wb.create_sheet("Pricing Details")
        costs_sheet = wb["Monthly Costs"]
        costs_sheet.append(["Month", "Food Costs", "Beverage Costs", "Chemical Costs", "Total Costs"])
        sales_sheet = wb["Sales Data"]
        sales_sheet.append(["Date", "Category", "Item", "Quantity Sold", "Revenue"])
        pricing_sheet = wb["Pricing Details"]
        pricing_sheet.append(["Item Name", "Category", "Supplier", "Unit Price"])
        wb.save(CENTRAL_SPREADSHEET)
    return "Central spreadsheet created or exists."

# Function to validate extracted data
def validate_extracted_data(data, expected_keys):
    errors = []
    for key in expected_keys:
        if key not in data or not data[key]:
            errors.append(f"Missing or empty field: {key}")
    return errors

# Function to update the spreadsheet with validation
def update_spreadsheet_with_validation(extracted_data, sheet_name):
    wb = load_workbook(CENTRAL_SPREADSHEET)
    error_sheet = wb.create_sheet("Error Log") if "Error Log" not in wb.sheetnames else wb["Error Log"]
    validation_errors = validate_extracted_data(extracted_data, expected_keys={"Pricing Details": ["Item Name", "Category"]})
    if validation_errors:
        error_sheet.append([datetime.now(), "; ".join(validation_errors), json.dumps(extracted_data)])
        wb.save(CENTRAL_SPREADSHEET)
        return "Error logged."
    sheet = wb[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    row = [extracted_data.get(header, "") for header in headers]
    sheet.append(row)
    wb.save(CENTRAL_SPREADSHEET)
    return "Data added."

# Main processing function
def process_and_save_files():
    create_central_spreadsheet()  # Ensure the central spreadsheet exists
    for filename in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.isfile(file_path) and allowed_file(filename):
            print(f"Processing file: {filename}")
            file_ext = filename.rsplit('.', 1)[1].lower()
            content = ""

            try:
                if file_ext == 'pdf':
                    content = process_pdfs(file_path)
                elif file_ext in ['png', 'jpeg', 'jpg']:
                    content = process_images(file_path)
                elif file_ext == 'xlsx':
                    content = process_excel(file_path)
                else:
                    print(f"Unsupported file type: {file_ext}")
                    continue

                # Define output .txt file path
                output_path = os.path.splitext(file_path)[0] + '.txt'

                # Save the extracted content to a .txt file
                with open(output_path, 'w', encoding='utf-8') as f:
                    if isinstance(content, str):
                        f.write(content)
                    elif isinstance(content, list):  # Assuming content from Excel is a list of dictionaries
                        for record in content:
                            f.write(json.dumps(record) + '\n')
                    else:
                        f.write(str(content))
                
                print(f"File processed and saved to: {output_path}")

                # Update the central spreadsheet
                if isinstance(content, str):  # Example logic for updating Pricing Details
                    update_spreadsheet_with_validation({"Item Name": filename, "Category": "General"}, "Pricing Details")
                elif isinstance(content, list):
                    for item in content:
                        update_spreadsheet_with_validation(item, "Pricing Details")

            except Exception as e:
                print(f"Error processing file {filename}: {e}")

# Run the processing function
if __name__ == "__main__":
    process_and_save_files()